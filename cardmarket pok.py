import camelot
import os
from openpyxl import load_workbook

# Camelot gives you a list of tables; each is like a pandas DataFrame
#table = tables[0].df   # DataFrame of the first table

#check if the two cards are the same
def check_equal(sheet, poke_row, row, col):
    flag=True
    if sheet.cell(row=row, column=col+2).value != poke_row[3]: #language
        flag=False
    if flag==True and sheet.cell(row=row, column=col+5).value != poke_row[4]: #condition
        flag=False
    if flag==True and sheet.cell(row=row, column=col+6).value != poke_row[5]: #set
        flag=False
    if flag==True and sheet.cell(row=row, column=col+7).value != poke_row[2]: #number card
        flag=False
    if flag==True and sheet.cell(row=row, column=col+8).value != poke_row[6]: #rarity
        flag=False
    if flag==True and sheet.cell(row=row, column=col+3).value != poke_row[7]: #price
        flag=False
    return flag


def insert_card(ws, poke_row, row, col):
    ws.cell(row=row, column=col+1, value="✔") #checker
    ws.cell(row=row, column=col+2, value=poke_row[3]) #language
    ws.cell(row=row, column=col+3, value=poke_row[7]) #price
    ws.cell(row=row, column=col+4, value=1) #quantity
    ws.cell(row=row, column=col+5, value=poke_row[4]) #condition
    ws.cell(row=row, column=col+6, value=poke_row[5]) #set
    ws.cell(row=row, column=col+7, value=poke_row[2]) #number card
    ws.cell(row=row, column=col+8, value=poke_row[6]) #rarity

def write_in_second_page(poke_row):
    second_ws=wb.worksheets[1] #second page
    row = 2
    while second_ws.cell(row=row, column=1).value is not None:
        flag=check_equal(second_ws, poke_row, row, 0)
        if flag==True:
            return
        else:
            row += 1
    second_ws.cell(row=row, column=1, value=poke_row[1]) #name
    second_ws.cell(row=row, column=2, value=poke_row[3]) #language
    second_ws.cell(row=row, column=3, value=poke_row[7]) #price
    second_ws.cell(row=row, column=4, value=1) #quantity
    second_ws.cell(row=row, column=5, value=poke_row[4]) #condition
    second_ws.cell(row=row, column=6, value=poke_row[5]) #set
    second_ws.cell(row=row, column=7, value=poke_row[2]) #number card
    second_ws.cell(row=row, column=8, value=poke_row[6]) #rarity           


cartella = r"C:\Users\Simigliani\Desktop\Programmi C\Progetti\cardmarket converter list\acquisti"



for file in os.listdir(cartella):
    if file.lower().endswith(".pdf"):
        percorso = os.path.join(cartella, file)

        # Extract tables from a PDF (first page, but you can set pages="all")
        tables = camelot.read_pdf(percorso, pages="all")

        # Open Excel file
        wb = load_workbook(r"C:\Users\Simigliani\Desktop\Programmi C\Progetti\cardmarket converter list\Check Pokemon.xlsx")
        ws = wb.active   # active page (can use wb["NomeFoglio"])

        #take name from pdf table
        for table in tables:
            df =table.df
            for poke_row in df.itertuples(index=False):
                first_element = poke_row[0]

                colonne_name = [2,13,24,35,46,57,68,79,90,101]

                for col in colonne_name:
                    for row in range(2, ws.max_row + 1): #start from row 2 until the end
                        cell=ws.cell(row=row, column=col)
                        if cell.value == poke_row[1]:
                            print(f"Found {poke_row[1]} at row {row}, column {col}")

                            if ws.cell(row=row, column=col+1).value != "✔":
                                insert_card(ws,poke_row, row, col)
                            else:
                                flag=check_equal(ws, poke_row, row, col)
                                if flag==False:
                                    #write card in the second page
                                    write_in_second_page(poke_row)



        #save file
        wb.save(r"C:\Users\Simigliani\Desktop\Programmi C\Progetti\cardmarket converter list\Check Pokemon.xlsx")
    

